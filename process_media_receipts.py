#!/usr/bin/env python3
"""
ZGM Media Receipt Processing System
=====================================
Processes media invoices/receipts from vendors and generates
FunctionPointe-compatible import CSVs.

Supported Naming Conventions:
  [Supplier] - [InvoiceNumber] [JobCode] [Description].pdf   (most common)
  [Supplier] - [InvoiceNumber] [Date] [JobCode] [Desc].pdf
  [Supplier] [InvoiceNumber] - [JobCode] [Description].pdf
  e.g., "CFAC - 1016060-3 DE-3032 Radio.pdf"
        "Netflix - CINV-5417518 Feb 5 2026 JAY-3313 Digital.pdf"
        "Netflix CINV-5414568 - EE 3270.pdf"

Output: FunctionPointe External Expense CSV
"""

import os
import csv
import json
import re
import shutil
import tempfile
import argparse
import pdfplumber
import requests
from datetime import datetime
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Paths ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(SCRIPT_DIR, ".env"))

RESOURCES_FOLDER       = os.path.join(SCRIPT_DIR, "resources")
SUPPLIER_CODES_FILE    = os.path.join(RESOURCES_FOLDER, "FP_Supplier_Codes.csv")

FP_API_BASE_URL        = "https://api-platform.functionpoint.com"
FP_API_TIMEOUT_SECONDS = 30
DROPBOX_API_BASE_URL   = "https://api.dropboxapi.com/2"
DROPBOX_CONTENT_URL    = "https://content.dropboxapi.com/2"
DROPBOX_TIMEOUT_SECONDS = 60

APP_TIMEZONE = os.getenv("APP_TIMEZONE", "America/Edmonton")
RUN_STARTED_AT = datetime.now(ZoneInfo(APP_TIMEZONE))
RUN_MONTH_FOLDER = RUN_STARTED_AT.strftime("%b %Y")
TIMESTAMP = RUN_STARTED_AT.strftime("%b_%Y_%d_%H%M%S")


def set_processing_root(root_folder):
    """Point all operational folders at a local processing root."""
    global INCOMING_FOLDER, PROCESSED_FOLDER, OUTPUT_FOLDER, ERROR_FOLDER
    global PROCESSED_BASE_FOLDER, OUTPUT_BASE_FOLDER
    global MANUAL_REVIEW_FOLDER, MANUAL_PO_FOLDER, MANUAL_MULTI_FOLDER
    global NAMING_ERRORS_FOLDER

    INCOMING_FOLDER        = os.path.join(root_folder, "Incoming")
    PROCESSED_BASE_FOLDER  = os.path.join(root_folder, "Processed")
    OUTPUT_BASE_FOLDER     = os.path.join(root_folder, "Output")
    PROCESSED_FOLDER       = os.path.join(PROCESSED_BASE_FOLDER, RUN_MONTH_FOLDER)
    OUTPUT_FOLDER          = os.path.join(OUTPUT_BASE_FOLDER, RUN_MONTH_FOLDER)
    ERROR_FOLDER           = os.path.join(root_folder, "Error")
    MANUAL_REVIEW_FOLDER   = os.path.join(root_folder, "Manual Review")
    MANUAL_PO_FOLDER       = os.path.join(root_folder, "Manual Enter - PO")
    MANUAL_MULTI_FOLDER    = os.path.join(root_folder, "Manual Enter - Multi-Job")
    NAMING_ERRORS_FOLDER   = os.path.join(root_folder, "Naming Errors")


set_processing_root(SCRIPT_DIR)

# ─── FP Import Defaults ───────────────────────────────────────────────────────
# These defaults can be adjusted to match your FP configuration.
FP_DEFAULTS = {
    "Payable_Account": "",        # e.g. "2000" — leave blank to fill in FP
    "Office":          "CGY",     # Calgary office code
    "Terms":           "Net 30",
    "Quantity":        "1",
    "Markup_Pct":      "0",
    "Billed":          "Yes",     # Bill to client
    "Override":        "",
    "Manually_Exported": "",
    "Start_New_Expense": "",
    "Tip":             "",
    "Discount":        "",
}

# Tax groups — Canadian vendors charge GST; US-origin digital services typically 0
TAX_GROUP_GST  = "GST"
TAX_GROUP_NONE = ""


class FunctionPointLookupError(RuntimeError):
    """Raised when a job or expense mapping cannot be resolved safely."""


class FunctionPointServiceError(RuntimeError):
    """Raised when Function Point authentication or connectivity fails."""

# ─── Supplier Mapping ─────────────────────────────────────────────────────────
# Maps keywords found in filenames / PDF text to FP supplier codes.
# Add new vendors here as needed.
SUPPLIER_MAP = {
    # Keyword (lowercase)           : (FP_Code,       Display_Name,                       Tax_Group)
    # ── Digital / Social ──────────────────────────────────────────────────────────────────────────
    "meta":                           ("Fac",           "Meta / Facebook",                  TAX_GROUP_NONE),
    "facebook":                       ("Fac",           "Meta / Facebook",                  TAX_GROUP_NONE),
    "netflix":                        ("Netflix",       "Netflix",                          TAX_GROUP_GST),
    "google":                         ("GoAdW",         "Google Ads",                       TAX_GROUP_NONE),
    "youtube":                        ("YT",            "YouTube",                          TAX_GROUP_NONE),
    "tiktok":                         ("Tik",           "TikTok",                           TAX_GROUP_NONE),
    "twitter":                        ("Twi",           "Twitter / X",                      TAX_GROUP_NONE),
    "spotify":                        ("Spo",           "Spotify",                          TAX_GROUP_NONE),
    "vistar":                         ("ViMed",         "Vistar Media",                     TAX_GROUP_GST),
    "nexthome":                       ("NexHom",        "NextHome",                         TAX_GROUP_GST),
    "campsite":                       ("CaGlInc",       "Campsite Global Inc.",              TAX_GROUP_GST),
    "ai digital":                     ("AIDig",         "AI Digital",                       TAX_GROUP_GST),
    "infinite gravity":               ("InGrDiMeLt",    "Infinite Gravity Digital",         TAX_GROUP_GST),
    "cineplex":                       ("CiDiMeInc",     "Cineplex Digital Media",           TAX_GROUP_GST),
    # ── Programmatic / Ad Tech ────────────────────────────────────────────────────────────────────
    "dandelion":                      ("DaInc",         "Dandelion Inc",                    TAX_GROUP_GST),
    "adcanada":                       ("AdMeInc",       "AdCanada Media Inc.",              TAX_GROUP_GST),
    "rec media":                      ("REMed",         "REC Media",                        TAX_GROUP_GST),
    # ── Out of Home ───────────────────────────────────────────────────────────────────────────────
    "pattison outdoor":               ("PaOuAdL",       "Pattison Outdoor Advertising LP",  TAX_GROUP_GST),
    "pattison":                       ("PaOuAdL",       "Pattison Outdoor Advertising LP",  TAX_GROUP_GST),
    "farwest outdoor":                ("FarWest",        "FarWest Outdoor",                  TAX_GROUP_GST),
    "farwest":                        ("FarWest",        "FarWest Outdoor",                  TAX_GROUP_GST),
    "oilers entertainment":           ("OiEnGro",       "Oilers Entertainment Group",       TAX_GROUP_GST),
    "oilers":                         ("OiEnGro",       "Oilers Entertainment Group",       TAX_GROUP_GST),
    # ── Television ────────────────────────────────────────────────────────────────────────────────
    "cbxt":                           ("CBXT",          "CBXT",                             TAX_GROUP_GST),
    "ctv":                            ("CTVC",          "CTV",                              TAX_GROUP_GST),
    "global television":              ("GT",            "Global Television",                TAX_GROUP_GST),
    # ── Radio ─────────────────────────────────────────────────────────────────────────────────────
    "cfac":                           ("CFAC",          "CFAC",                             TAX_GROUP_GST),
    "cfxl-fm":                        ("CFXLFM",        "CFXL-FM",                          TAX_GROUP_GST),
    "cfxl":                           ("CFXLFM",        "CFXL-FM",                          TAX_GROUP_GST),
    "cirk-fm":                        ("CIRKFM",        "CIRK-FM",                          TAX_GROUP_GST),
    "cirk":                           ("CIRKFM",        "CIRK-FM",                          TAX_GROUP_GST),
    "ckmp-fm":                        ("CKMPFM",        "CKMP-FM",                          TAX_GROUP_GST),
    "ckmp":                           ("CKMPFM",        "CKMP-FM",                          TAX_GROUP_GST),
    "ckra-fm":                        ("CKRAFM",        "CKRA-FM",                          TAX_GROUP_GST),
    "ckra":                           ("CKRAFM",        "CKRA-FM",                          TAX_GROUP_GST),
    "bell media":                     ("BMRGPC",        "Bell Media",                       TAX_GROUP_GST),
    "corus":                          ("CSI",           "Corus Sales Inc",                  TAX_GROUP_GST),
    "rogers digital":                 ("RoDiMed",       "Rogers Digital Media",             TAX_GROUP_GST),
    # ── Print / Publishing ────────────────────────────────────────────────────────────────────────
    "paper leaf":                     ("PaLea",         "Paper Leaf",                       TAX_GROUP_GST),
}

# ─── Job Code Pattern ─────────────────────────────────────────────────────────
# ZGM job codes are: 2-6 uppercase letters + dash + exactly 4 digits
# e.g. DCC-3074, DE-3237, EE-3270, AIRB-3002, COVH-3173
# Also handles flexible spacing: "DCC 3074", "EE - 3270", "FOR -3412"
# Intentionally excludes: 5-digit numbers (BOX-12814, CA-94025),
# 3-digit numbers (CAD-198), and known non-job prefixes (INV-, RT-, TQ-)
JOB_CODE_PATTERN = re.compile(
    r'(?<![A-Z0-9])'
    r'([A-Z]{2,6}(?:[ \t]*-[ \t]*|[ \t]+)[0-9]{4})'
    r'(?![0-9])'
)

# Prefixes that are invoice/reference numbers or media terms — NOT job codes
NON_JOB_PREFIXES = {'INV', 'REF', 'PO', 'RT', 'TQ', 'TAX', 'REC', 'CAD', 'USD', 'CPM', 'CPC', 'CPA', 'CTR'}

# Words that appear before an invoice number but are NOT the invoice number themselves
INVOICE_LABEL_WORDS = {'inv', 'no', 'no.', 's', 'ref', 'invoice'}

# Month name abbreviations (used to skip date tokens during invoice parsing)
MONTH_ABBREVS = {'jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec',
                 'january','february','march','april','june','july','august',
                 'september','october','november','december'}

def normalize_job_code(code):
    """Normalize flexible separators and spaces to PREFIX-1234."""
    return re.sub(
        r'[ \t]*-[ \t]*|[ \t]+', '-', code.strip().upper()
    )


def _is_valid_job_code(code):
    """
    Return True if code looks like a ZGM job code (LETTERS-4DIGITS).
    Filters out invoice numbers, media terms, and other false positives.
    """
    parts = code.split('-')
    if len(parts) != 2:
        return False
    prefix, number = parts
    if not re.match(r'^\d{4}$', number):   # must be exactly 4 digits
        return False
    if prefix in NON_JOB_PREFIXES:
        return False
    if not re.match(r'^[A-Z]{2,6}$', prefix):
        return False
    return True


def parse_filename_components(filename):
    """
    Parse filename into (supplier_hint, invoice_num, service_group).

    Handles all ZGM naming variants:
      Supplier - Invoice JobCode ServiceGroup.pdf
      Supplier - Invoice Date JobCode ServiceGroup.pdf
      Supplier Invoice - JobCode ServiceGroup.pdf
      Supplier Invoice JobCode ServiceGroup.pdf   (no separator)

    Uses the first valid job code as the anchor point.
    Invoice number = first digit-containing token after the supplier name.
    Service group  = everything after the last job code.
    """
    name = os.path.splitext(filename)[0]

    # ── Find job code positions ────────────────────────────────────────────────
    first_job_match = None
    last_job_match  = None
    for m in JOB_CODE_PATTERN.finditer(name):
        code = normalize_job_code(m.group(1))
        if _is_valid_job_code(code):
            if first_job_match is None:
                first_job_match = m
            last_job_match = m

    if first_job_match is None:
        return name.strip(), "N/A", ""

    pre_job     = name[:first_job_match.start()].strip().strip('-').strip()
    post_job    = name[last_job_match.end():].strip().strip('-').strip()
    service_group = post_job  # everything after the last job code

    # ── Parse supplier + invoice from pre_job ─────────────────────────────────
    supplier_hint = ""
    invoice_num   = "N/A"

    def first_invoice_token(token_list):
        """Return first invoice token, including any label prefix (e.g. 'inv 42575').
        Label words (inv, no, ref…) are kept when they immediately precede a digit token."""
        pending_label = None
        for tok in token_list:
            tl = tok.lower()
            if tl in INVOICE_LABEL_WORDS:
                pending_label = tok   # may be a prefix — wait for the next digit token
                continue
            if tl in MONTH_ABBREVS:
                pending_label = None  # month resets label context
                continue
            if re.search(r'\d', tok):
                return f"{pending_label} {tok}" if pending_label else tok
            pending_label = None      # non-label, non-digit word resets context
        return "N/A"

    if ' - ' in pre_job:
        # Split on ALL ' - ' occurrences
        sections = [s.strip() for s in pre_job.split(' - ') if s.strip()]
        supplier_hint = sections[0]  # first section = supplier

        # Scan remaining sections for the invoice number
        for section in sections[1:]:
            tok = first_invoice_token(section.split())
            if tok != "N/A":
                invoice_num = tok
                break
    else:
        # No ' - ': supplier = all words before the first digit-containing token
        # Track pending label so "inv 42575" stays together
        tokens = pre_job.split()
        invoice_idx  = None
        label_idx    = None
        pending_label = None
        for i, tok in enumerate(tokens):
            tl = tok.lower()
            if tl in INVOICE_LABEL_WORDS:
                pending_label = tok
                label_idx = i
                continue
            if tl in MONTH_ABBREVS:
                pending_label = None
                label_idx = None
                continue
            if re.search(r'\d', tok):
                invoice_idx = i
                break
            pending_label = None
            label_idx = None

        if invoice_idx is not None:
            start = label_idx if (label_idx is not None and label_idx == invoice_idx - 1) else invoice_idx
            supplier_hint = ' '.join(tokens[:start])
            invoice_num   = f"{tokens[label_idx]} {tokens[invoice_idx]}" if start == label_idx else tokens[invoice_idx]
        else:
            supplier_hint = pre_job

    return supplier_hint.strip(), invoice_num, service_group.strip()


# ─── Supplier Detection ───────────────────────────────────────────────────────

def load_supplier_codes():
    """Load FP supplier codes from CSV into a dict: name_lower -> code."""
    codes = {}
    if os.path.exists(SUPPLIER_CODES_FILE):
        with open(SUPPLIER_CODES_FILE, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                codes[row['Supplier_Name'].lower()] = row['FP_Code']
    return codes


def detect_supplier(filename, text, fp_codes):
    """
    Detect supplier from filename and PDF text.
    Returns (fp_code, display_name, tax_group, confidence)
    confidence: 'HIGH', 'MEDIUM', 'LOW'
    """
    filename_lower = filename.lower()
    text_lower = text.lower() if text else ""

    # Check known mappings (filename first, then text)
    for keyword, (fp_code, display_name, tax_group) in SUPPLIER_MAP.items():
        if keyword in filename_lower:
            return fp_code, display_name, tax_group, "HIGH"

    for keyword, (fp_code, display_name, tax_group) in SUPPLIER_MAP.items():
        if keyword in text_lower:
            return fp_code, display_name, tax_group, "LOW"

    # Fallback: extract supplier name from filename (everything before first " - ")
    # This takes priority over fuzzy FP text matching to avoid wrong matches
    name_no_ext = os.path.splitext(filename)[0]
    if ' - ' in name_no_ext:
        extracted = name_no_ext.split(' - ')[0].strip()
        if extracted:
            exact_code = fp_codes.get(extracted.lower())
            if exact_code:
                return exact_code, extracted, TAX_GROUP_NONE, "HIGH"
            return extracted, extracted, TAX_GROUP_NONE, "LOW"

    # Last resort: try matching against full FP supplier list
    for name_lower, code in fp_codes.items():
        if name_lower[:8] in filename_lower or name_lower[:8] in text_lower:
            return code, name_lower.title(), TAX_GROUP_NONE, "LOW"

    return "UNKNOWN", "Unknown Supplier", TAX_GROUP_NONE, "LOW"


# ─── Invoice Number Extraction ────────────────────────────────────────────────

def extract_invoice_number(filename, text):
    """
    Extract invoice/reference number from filename using parse_filename_components.
    Falls back to PDF text if filename parsing fails.

    Invoice numbers may contain dashes with NO spaces (e.g. 1015821-2, E-580582).
    """
    _, invoice_num, _ = parse_filename_components(filename)

    if invoice_num != "N/A":
        return invoice_num

    # Fallback for "Multiple" filenames (e.g. Dandelion): scan the filename directly
    # for common invoice number tokens like INV-13434 or 12345-6
    name_no_ext = os.path.splitext(filename)[0]
    if 'multiple' in name_no_ext.lower():
        m = re.search(r'\b((?:INV|CINV|REF|PO)-?\d[\w\-]*\d|\d[\d\-]{2,}\d)\b', name_no_ext, re.IGNORECASE)
        if m:
            cand = m.group(1)
            # Make sure it's not a date year or short noise
            if not re.match(r'^(19|20)\d{2}$', cand):
                return cand

    # Fallback: scan PDF text for common invoice number patterns
    if text:
        for pat in [
            r'Invoice\s*(?:ID|#|Number|No\.?)[\s:]+([A-Z0-9][A-Z0-9\-]+)',
            r'Reference\s*Number:\s+([A-Z0-9]+)',
            r'\bINVOICE\s+([A-Z]{2,4}-\d{4,6})\b',
        ]:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                return m.group(1).strip()

    return "N/A"


def extract_service_group(filename):
    """
    Extract the service group from the filename.
    Service group = everything after the last job code
    (e.g. 'Radio', 'Out Of Home', 'Digital - OL1', 'Video Production').
    """
    _, _, service_group = parse_filename_components(filename)
    return service_group


# ─── Date Extraction ──────────────────────────────────────────────────────────

def extract_date(filename, text):
    """Extract invoice/expense date from PDF text."""
    if not text:
        return "N/A", "LOW"

    # Strip lines whose labels indicate a date range or period — not the invoice date.
    # "Invoice Period", "Flight Dates", "Air Dates" etc. are schedule metadata, not billing dates.
    IGNORE_DATE_LABELS = [
        'invoice period', 'invoice month', 'flight dates', 'flight date',
        'air dates', 'air date', 'broadcast period', 'broadcast dates',
        'broadcast date', 'schedule period', 'billing period',
    ]
    cleaned_lines = []
    for line in text.split('\n'):
        if any(label in line.lower() for label in IGNORE_DATE_LABELS):
            cleaned_lines.append('')   # blank the line so its dates are invisible
        else:
            cleaned_lines.append(line)
    text = '\n'.join(cleaned_lines)

    # Common date patterns in media invoices.
    # Each tuple: (regex, list_of_strptime_formats)
    # NOTE: the FIRST capture group in each regex must capture the FULL date string.
    # Ordered from most specific / highest confidence to most general.
    patterns = [
        # Labelled full-month-name date: "Invoice Date: January 5, 2026"
        (r'(?:Invoice\s+Date|Date\s+Issued|Issue\s+Date|Date\s+of\s+[Ii]ssue|'
         r'Billed\s+On|Bill\s+Date|Invoice/Payment\s+Date)[\s:/]+(\w+\s+\d{1,2},?\s+\d{4})',
         ['%B %d, %Y', '%b %d, %Y', '%B %d %Y', '%b %d %Y']),

        # Labelled ISO date: "Invoice Date 2025-12-31" / "Date: 2025-12-31"
        (r'(?:Invoice\s+Date|Date|DATE)[\s:/]+(\d{4}-\d{2}-\d{2})',
         ['%Y-%m-%d']),

        # Labelled 4-digit-year numeric: "Invoice Date: 01/05/2026" / "Transaction date: 3/12/2026"
        (r'(?:Invoice\s+Date|Transaction\s+[Dd]ate|Date|DATE)[\s:/]+(\d{1,2}/\d{1,2}/\d{4})',
         ['%m/%d/%Y', '%d/%m/%Y']),

        # Labelled 2-digit-year numeric: "Invoice Date 10/26/25"  (broadcast standard)
        # Python %y: 00-68 → 2000-2068, 69-99 → 1969-1999  (safe for current invoices)
        (r'(?:Invoice\s+Date|Date\s+Issued|Issue\s+Date)[\s:/]+(\d{2}/\d{2}/\d{2})\b',
         ['%m/%d/%y']),

        # DD-MON-YYYY: "26-OCT-2025"  (CBC / CBXT format)
        (r'\b(\d{1,2}-(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-\d{4})\b',
         ['%d-%b-%Y']),

        # Standalone full month name: "January 1, 2026" / "January 1 2026"
        (r'\b((?:January|February|March|April|May|June|July|August|September|'
         r'October|November|December)\s+\d{1,2},?\s+\d{4})\b',
         ['%B %d, %Y', '%B %d %Y']),

        # Standalone abbreviated month: "Jan 1, 2026"
        (r'\b((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\.?\s+\d{1,2},?\s+\d{4})\b',
         ['%b %d, %Y', '%b. %d, %Y', '%b %d %Y']),

        # Day-first abbreviated month: "19 Dec 2025"
        (r'\b(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\.?\s+\d{4})\b',
         ['%d %b %Y', '%d %b. %Y']),

        # Compressed day+month+year (no spaces — PDFs where text runs together): "19Dec2025"
        # Also catches "InvoiceDate" label run into the date on same token
        (r'\b(\d{1,2}(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\d{4})\b',
         ['%d%b%Y']),

        # Standalone numeric 4-digit year: "12/31/2025" or "3/12/2026"
        (r'\b(\d{1,2}/\d{1,2}/\d{4})\b',
         ['%m/%d/%Y', '%d/%m/%Y']),

        # ISO date standalone: "2025-12-31"
        (r'\b(\d{4}-\d{2}-\d{2})\b',
         ['%Y-%m-%d']),
    ]

    for pat, fmts in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            raw = m.group(1).strip()
            for fmt in fmts:
                try:
                    dt = datetime.strptime(raw, fmt)
                    return dt.strftime('%m-%d-%Y'), "HIGH"
                except ValueError:
                    continue
            return raw, "MEDIUM"

    return "N/A", "LOW"


# ─── Amount Extraction ────────────────────────────────────────────────────────

def extract_amount(text, supplier_key=""):
    """
    Extract total amount (before tax) and tax amount from invoice.
    Returns (subtotal, tax_amount, currency, confidence)
    """
    if not text:
        return "N/A", "N/A", "CAD", "LOW"

    currency = "CAD"  # default
    # Detect currency — check explicit indicators first, then text presence
    # Priority order: Payment Currency field > total line currency > document text
    if re.search(r'Payment Currency:\s*CAD', text):
        currency = "CAD"
    elif re.search(r'Payment Currency:\s*USD', text):
        currency = "USD"
    elif re.search(r'(?:TOTAL|Balance Due|Subtotal)\s*\n?\s*CAD', text, re.IGNORECASE):
        currency = "CAD"
    elif re.search(r'^CAD\s+[\d,]+', text, re.MULTILINE):
        currency = "CAD"
    elif re.search(r'Payable in Canadian', text, re.IGNORECASE):
        currency = "CAD"
    elif re.search(r'Canadian Dollars\s*\(CAD\)', text, re.IGNORECASE):
        currency = "CAD"
    elif re.search(r'CA\$|CDN\$|\bCAD\b', text):
        currency = "CAD"
    elif re.search(r'US\$|\bUSD\b', text):
        currency = "USD"

    subtotal = "N/A"
    tax_amount = "N/A"

    # ── Meta / Facebook specific ──────────────────────────────────────────
    if 'meta' in supplier_key or 'facebook' in supplier_key:
        # "CA$1,040.00 CAD" or "Subtotal: CA$1,040.00 CAD"
        m = re.search(r'Subtotal:\s*CA?\$?([\d,]+\.?\d*)\s*CAD', text)
        if m:
            subtotal = m.group(1).replace(',', '')
        else:
            m = re.search(r'Reference Number:\s+\w+\s+CA?\$?([\d,]+\.?\d*)', text)
            if m:
                subtotal = m.group(1).replace(',', '')
        # GST/HST for Meta is typically 0
        m_tax = re.search(r'GST/HST:\s*CAD\s*([\d,]+\.?\d*)', text)
        tax_amount = m_tax.group(1).replace(',', '') if m_tax else "0.00"
        return subtotal, tax_amount, currency, "HIGH"

    # ── Netflix specific ──────────────────────────────────────────────────
    if 'netflix' in supplier_key:
        # Netflix format: "Subtotal $4,584.18 \n GST 5% $229.21 \n TOTAL $4,813.39"
        m_sub = re.search(r'Subtotal\s+\$?([\d,]+\.?\d*)', text, re.IGNORECASE)
        if m_sub:
            subtotal = m_sub.group(1).replace(',', '')
        else:
            m_total = re.search(r'TOTAL\s+\$?([\d,]+\.?\d*)', text)
            if m_total:
                subtotal = m_total.group(1).replace(',', '')
        m_tax = re.search(r'GST\s+5%\s+\$?([\d,]+\.?\d*)', text, re.IGNORECASE)
        if not m_tax:
            m_tax = re.search(r'GST/HST\s+\$?([\d,]+\.?\d*)', text, re.IGNORECASE)
        tax_amount = m_tax.group(1).replace(',', '') if m_tax else "N/A"
        return subtotal, tax_amount, currency, "HIGH"

    # ── Dandelion / DV360 ─────────────────────────────────────────────────
    if 'dandelion' in supplier_key:
        # Dandelion shows SUBTOTAL then GST then TOTAL on separate lines
        m_sub = re.search(r'SUBTOTAL\s+([\d,]+\.?\d*)', text, re.IGNORECASE)
        if m_sub:
            subtotal = m_sub.group(1).replace(',', '')
        m_tax = re.search(r'GST/HST\s*@\s*5%\s+([\d,]+\.?\d*)', text, re.IGNORECASE)
        if not m_tax:
            m_tax = re.search(r'GST\s+([\d,]+\.?\d*)', text, re.IGNORECASE)
        tax_amount = m_tax.group(1).replace(',', '') if m_tax else "N/A"
        return subtotal, tax_amount, currency, "HIGH"

    # ── Generic patterns ──────────────────────────────────────────────────
    # Priority 1a: "Subtotal before taxes $X" — same line (REC Media when amount is adjacent)
    m = re.search(r'Subtotal before taxes\s*:?\s*\$?([\d,]+\.\d{2})', text, re.IGNORECASE)
    if m:
        subtotal = m.group(1).replace(',', '')

    # Priority 1b: "Subtotal before taxes" label present but amount is elsewhere on the page
    # → compute as Total Amount - Total Taxes (reliable for REC Media invoice layout)
    if subtotal == "N/A" and re.search(r'Subtotal before taxes', text, re.IGNORECASE):
        m_tot = re.search(r'Total\s+amount\s*\$?([\d,]+\.\d{2})', text, re.IGNORECASE)
        m_tax = re.search(r'Total\s+taxes?\s*\$?([\d,]+\.\d{2})', text, re.IGNORECASE)
        if m_tot and m_tax:
            try:
                subtotal = f"{float(m_tot.group(1).replace(',','')) - float(m_tax.group(1).replace(',','')):.2f}"
            except ValueError:
                pass

    # Priority 2: plain Subtotal / Sub Total / Sub-Total
    if subtotal == "N/A":
        for label in ['Subtotal', 'Sub Total', 'Sub-Total']:
            m = re.search(rf'{label}\s*:?\s*(?:CA\$|CDN\$|\$)?([\d,]+\.?\d*)', text, re.IGNORECASE)
            if m:
                subtotal = m.group(1).replace(',', '')
                break

    # Priority 3: "Net Total" — CFAC/Rogers broadcast format (distinct from "Net 30" payment terms)
    if subtotal == "N/A":
        m = re.search(r'\bNet\s+Total\b\s*\$?([\d,]+\.?\d*)', text, re.IGNORECASE)
        if m:
            subtotal = m.group(1).replace(',', '')

    # Priority 4: bare NET with $ — pre-commission pre-tax line (CBXT/CBC format)
    # Require $ to avoid matching "NET 30 DAYS" payment terms
    if subtotal == "N/A":
        m = re.search(r'\bNET\b\s*\$\s*([\d,]+\.?\d*)', text, re.IGNORECASE)
        if m:
            subtotal = m.group(1).replace(',', '')

    # Priority 5: plain Total (not Amount Due / Total Due / Total Tax lines)
    if subtotal == "N/A":
        m = re.search(r'\bTotal(?!\s+Amount)(?!\s+Due)(?!\s+Tax)\s+(?:CAD|CA\$|CDN\$|\$)?([\d,]+\.?\d*)', text, re.IGNORECASE)
        if m:
            subtotal = m.group(1).replace(',', '')

    # Priority 6: Amount Due / Balance Due — last resort, these include tax
    if subtotal == "N/A":
        for label in ['Total Amount Due', 'Amount Due', 'Balance Due', 'Total Due']:
            # Also handles bilingual labels like "TOTAL AMOUNT DUE / MONTANT TOTAL"
            m = re.search(rf'{label}[\s/A-ZÀ-Ÿ]*:?\s*(?:CAD|CA\$|CDN\$|\$)?\s*([\d,]+\.?\d*)', text, re.IGNORECASE)
            if m:
                subtotal = m.group(1).replace(',', '')
                break

    # GST/Tax extraction
    for gst_pat in [
        r'VAT/Tax\s*\([^)]*\)\s*(?:CA\$|CDN\$|\$)?([\d,]+\.?\d*)',  # Reddit: "VAT/Tax (5.00%) CA$42.57"
        r'GST\s+\(.*?\)\s+GST\s+\$?([\d,]+\.?\d*)',
        r'GST\s+\$?([\d,]+\.?\d*)',
        r'GST/HST\s+\$?([\d,]+\.?\d*)',
        r'Tax\s+\$?([\d,]+\.?\d*)',
    ]:
        m_tax = re.search(gst_pat, text, re.IGNORECASE)
        if m_tax:
            tax_amount = m_tax.group(1).replace(',', '')
            break

    confidence = "HIGH" if subtotal != "N/A" else "LOW"
    return subtotal, tax_amount, currency, confidence


# ─── Job Code Extraction ──────────────────────────────────────────────────────

def extract_job_codes(filename, text):
    """
    Extract ZGM job codes from filename and PDF text.
    Returns (job_codes_list, source)

    Strategy:
    - Filename job codes are the source of truth.
    - PDF job codes are added ONLY if the filename had none (e.g. Dandelion multi-job).
    - This prevents false multi-job flags from address codes, IO names, etc. in PDFs.
    """
    filename_codes = set()
    name_no_ext = os.path.splitext(filename)[0]
    for m in JOB_CODE_PATTERN.finditer(name_no_ext):
        code = normalize_job_code(m.group(1))
        if _is_valid_job_code(code):
            filename_codes.add(code)

    if filename_codes:
        return sorted(filename_codes), "filename"

    # No job codes in filename — scan PDF (e.g. Dandelion multi-job invoices)
    pdf_codes = set()
    if text:
        for m in JOB_CODE_PATTERN.finditer(text):
            code = normalize_job_code(m.group(1))
            if _is_valid_job_code(code):
                pdf_codes.add(code)

    return sorted(pdf_codes), "pdf"


# ─── Function Point Job / Expense Lookup ─────────────────────────────────────

def _normalized_match_value(value):
    """Normalize a name/code for case-insensitive supplier matching."""
    return re.sub(r'[^a-z0-9]', '', str(value or '').lower())


def get_function_point_job(job_number, session, cache):
    """
    Retrieve a full Function Point docket (job) by its visible job number.

    Function Point requires two calls:
      1. GET /dockets?number={job_number} to obtain the internal docketid.
      2. GET /dockets/{docketid} to obtain estimates, phases and services.
    """
    job_number = int(job_number)
    if job_number in cache:
        return cache[job_number]

    try:
        response = session.get(
            f"{FP_API_BASE_URL}/dockets",
            params={"number": job_number},
            timeout=FP_API_TIMEOUT_SECONDS,
        )
        response.raise_for_status()
        payload = response.json()
    except (requests.RequestException, ValueError) as exc:
        raise FunctionPointServiceError(
            f"Function Point job search failed for {job_number}: {exc}"
        ) from exc

    # application/json returns a list. JSON-LD responses wrap it in hydra:member.
    if isinstance(payload, dict):
        matches = payload.get("hydra:member", [])
    elif isinstance(payload, list):
        matches = payload
    else:
        matches = []

    exact_matches = [
        item for item in matches
        if str(item.get("number", "")) == str(job_number)
    ]
    if not exact_matches:
        raise FunctionPointLookupError(
            f"No Function Point job found for number {job_number}"
        )
    if len(exact_matches) > 1:
        raise FunctionPointLookupError(
            f"Multiple Function Point jobs found for number {job_number}"
        )

    docket_id = exact_matches[0].get("docketid")
    if not docket_id:
        raise FunctionPointLookupError(
            f"Function Point job {job_number} has no docketid"
        )

    try:
        response = session.get(
            f"{FP_API_BASE_URL}/dockets/{docket_id}",
            timeout=FP_API_TIMEOUT_SECONDS,
        )
        response.raise_for_status()
        job = response.json()
    except (requests.RequestException, ValueError) as exc:
        raise FunctionPointServiceError(
            f"Function Point job detail failed for {job_number}: {exc}"
        ) from exc

    cache[job_number] = job
    return job


def _expense_candidates(job):
    """Return unique external-expense options and their parent service groups."""
    candidates = {}
    for estimate in job.get("estimates", []):
        for phase in estimate.get("estimatePhases", []):
            service_group = (phase.get("name") or "").strip()
            for service in phase.get("estimateServices", []):
                expense = service.get("externalexpense")
                if not isinstance(expense, dict):
                    continue

                candidate = {
                    "service_group": service_group,
                    "expense_type": (expense.get("name") or "").strip(),
                    "expense_type_id": expense.get("externalexpenseid"),
                    "expense_type_code": (expense.get("code") or "").strip(),
                    "service_name": (service.get("name") or "").strip(),
                    "service_description": (service.get("description") or "").strip(),
                }
                # Both values are required by the FP import. Incomplete API
                # options are not safe candidates for automatic processing.
                if not candidate["service_group"] or not candidate["expense_type"]:
                    continue

                key = (
                    candidate["service_group"],
                    candidate["expense_type_id"],
                    candidate["expense_type_code"],
                    candidate["expense_type"],
                )
                candidates[key] = candidate

    return list(candidates.values())


def match_supplier_to_function_point_expense(job, fp_code, supplier_display):
    """
    Match the detected invoice supplier to a job's external-expense options.

    Code matches are strongest, followed by exact names and known supplier aliases.
    A tie is rejected rather than selecting an arbitrary expense type.
    """
    candidates = _expense_candidates(job)
    if not candidates:
        raise FunctionPointLookupError(
            "Function Point job has no external expense options"
        )

    supplier_code = _normalized_match_value(fp_code)
    supplier_name = _normalized_match_value(supplier_display)
    aliases = set()
    for keyword, (mapped_code, mapped_name, _) in SUPPLIER_MAP.items():
        if (_normalized_match_value(mapped_code) == supplier_code or
                _normalized_match_value(mapped_name) == supplier_name):
            aliases.add(_normalized_match_value(keyword))
            aliases.add(_normalized_match_value(mapped_name))
    aliases.discard("")

    scored = []
    for candidate in candidates:
        expense_code = _normalized_match_value(candidate["expense_type_code"])
        expense_name = _normalized_match_value(candidate["expense_type"])
        score = 0

        if supplier_code and expense_code == supplier_code:
            score = max(score, 100)
        if supplier_name and expense_name == supplier_name:
            score = max(score, 95)
        if supplier_name and expense_name and (
                supplier_name in expense_name or expense_name in supplier_name):
            score = max(score, 80)
        for alias in aliases:
            if expense_name == alias:
                score = max(score, 90)
            elif alias in expense_name or expense_name in alias:
                score = max(score, 70)

        if score:
            scored.append((score, candidate))

    if not scored:
        raise FunctionPointLookupError(
            f"No Function Point expense matches supplier {supplier_display} [{fp_code}]"
        )

    best_score = max(score for score, _ in scored)
    best_matches = [candidate for score, candidate in scored if score == best_score]
    if len(best_matches) > 1:
        choices = ", ".join(
            f"{item['service_group']} / {item['expense_type']}"
            for item in best_matches
        )
        raise FunctionPointLookupError(
            f"Ambiguous Function Point expense match for {supplier_display}: {choices}"
        )

    return best_matches[0]


# ─── Description Builder ──────────────────────────────────────────────────────

def build_description(supplier_name, invoice_num, job_codes, text=""):
    """Build a concise description for the FP expense line."""
    job_str = ", ".join(job_codes) if job_codes else "See invoice"
    desc = f"{supplier_name} - {invoice_str(invoice_num)} - Job: {job_str}"
    return desc[:200]  # FP likely has a character limit


def invoice_str(inv):
    return inv if inv != "N/A" else "See invoice"


# ─── Validate Filename ────────────────────────────────────────────────────────

def validate_filename(filename):
    """
    Validate that a filename is processable.
    Only flags as a naming error if we truly cannot extract a job code —
    the naming convention is flexible (dash after supplier is optional, etc.).
    Returns (is_valid, issues_list)
    """
    issues = []

    if not filename.lower().endswith('.pdf'):
        issues.append("Not a PDF file")
        return False, issues

    name_no_ext = os.path.splitext(filename)[0]

    # Special case: "Multiple" in filename is an intentional multi-job placeholder
    # (e.g. Dandelion invoices that span many jobs) — treat as valid, job codes come from PDF
    if 'multiple' in name_no_ext.lower():
        return True, []

    # Only hard requirement: must have at least one valid job code in the filename
    has_job = any(
        _is_valid_job_code(normalize_job_code(m.group(1)))
        for m in JOB_CODE_PATTERN.finditer(name_no_ext)
    )
    if not has_job:
        issues.append("No job code found in filename (expected e.g. DCC-3074, AIRB-3002)")

    return len(issues) == 0, issues


# ─── PDF Text Extraction ──────────────────────────────────────────────────────

def extract_pdf_text(pdf_path):
    """Extract all text from a PDF file."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
            return text
    except Exception as e:
        return None


# ─── Move File ────────────────────────────────────────────────────────────────

def move_file(src, dest_folder, prefix=""):
    """Move a file to dest_folder, handling duplicates."""
    filename = os.path.basename(src)
    if prefix:
        filename = f"{prefix}_{filename}"
    dst = os.path.join(dest_folder, filename)
    if os.path.exists(dst):
        base, ext = os.path.splitext(filename)
        counter = 1
        while os.path.exists(dst):
            dst = os.path.join(dest_folder, f"{base}_{counter}{ext}")
            counter += 1
    shutil.move(src, dst)
    return dst


class DropboxIntegrationError(RuntimeError):
    """Raised when Dropbox authentication or file operations fail."""


DROPBOX_ROUTE_FOLDERS = [
    "Processed",
    "Error",
    "Manual Review",
    "Manual Enter - PO",
    "Manual Enter - Multi-Job",
    "Naming Errors",
]


def _dropbox_remote_path(root, *parts):
    """Build an absolute Dropbox path using forward slashes."""
    clean_root = "/" + str(root or "").strip().strip("/").strip()
    clean_parts = [str(part).strip("/") for part in parts if str(part).strip("/")]
    return "/".join([clean_root.rstrip("/")] + clean_parts)


def _dropbox_error(response, action):
    if response.ok:
        return
    detail = response.text[:1000]
    raise DropboxIntegrationError(
        f"Dropbox {action} failed ({response.status_code}): {detail}"
    )


def get_dropbox_access_token():
    """Exchange the stored refresh token for a temporary Dropbox access token."""
    required = [
        "DROPBOX_APP_KEY",
        "DROPBOX_APP_SECRET",
        "DROPBOX_REFRESH_TOKEN",
    ]
    missing = [name for name in required if not os.getenv(name, "").strip()]
    if missing:
        raise DropboxIntegrationError(
            f"Missing Dropbox configuration: {', '.join(missing)}"
        )

    response = requests.post(
        "https://api.dropboxapi.com/oauth2/token",
        data={
            "grant_type": "refresh_token",
            "refresh_token": os.environ["DROPBOX_REFRESH_TOKEN"].strip(),
        },
        auth=(
            os.environ["DROPBOX_APP_KEY"].strip(),
            os.environ["DROPBOX_APP_SECRET"].strip(),
        ),
        timeout=DROPBOX_TIMEOUT_SECONDS,
    )
    _dropbox_error(response, "token refresh")
    token = response.json().get("access_token")
    if not token:
        raise DropboxIntegrationError(
            "Dropbox token refresh returned no access_token"
        )
    return token


def _dropbox_api_headers(access_token):
    return {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }


def dropbox_list_folder(access_token, remote_folder):
    """Return every direct entry in a Dropbox folder, following pagination."""
    headers = _dropbox_api_headers(access_token)
    response = requests.post(
        f"{DROPBOX_API_BASE_URL}/files/list_folder",
        headers=headers,
        json={
            "path": remote_folder,
            "recursive": False,
            "include_deleted": False,
        },
        timeout=DROPBOX_TIMEOUT_SECONDS,
    )
    _dropbox_error(response, f"list {remote_folder}")
    payload = response.json()
    entries = list(payload.get("entries", []))

    while payload.get("has_more"):
        response = requests.post(
            f"{DROPBOX_API_BASE_URL}/files/list_folder/continue",
            headers=headers,
            json={"cursor": payload["cursor"]},
            timeout=DROPBOX_TIMEOUT_SECONDS,
        )
        _dropbox_error(response, f"continue listing {remote_folder}")
        payload = response.json()
        entries.extend(payload.get("entries", []))

    return entries


def dropbox_download_file(access_token, remote_path, local_path):
    """Download one Dropbox file to a local staging path."""
    response = requests.post(
        f"{DROPBOX_CONTENT_URL}/files/download",
        headers={
            "Authorization": f"Bearer {access_token}",
            "Dropbox-API-Arg": json.dumps({"path": remote_path}),
        },
        timeout=DROPBOX_TIMEOUT_SECONDS,
    )
    _dropbox_error(response, f"download {remote_path}")
    with open(local_path, "wb") as output_file:
        output_file.write(response.content)


def dropbox_upload_file(access_token, local_path, remote_path):
    """Upload a generated file without overwriting an existing Dropbox file."""
    with open(local_path, "rb") as input_file:
        response = requests.post(
            f"{DROPBOX_CONTENT_URL}/files/upload",
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/octet-stream",
                "Dropbox-API-Arg": json.dumps({
                    "path": remote_path,
                    "mode": "add",
                    "autorename": True,
                    "mute": False,
                }),
            },
            data=input_file,
            timeout=DROPBOX_TIMEOUT_SECONDS,
        )
    _dropbox_error(response, f"upload {remote_path}")
    return response.json()


def dropbox_ensure_folder(access_token, remote_folder):
    """Create a Dropbox folder when missing; accept an existing folder."""
    response = requests.post(
        f"{DROPBOX_API_BASE_URL}/files/create_folder_v2",
        headers=_dropbox_api_headers(access_token),
        json={"path": remote_folder, "autorename": False},
        timeout=DROPBOX_TIMEOUT_SECONDS,
    )
    if response.ok:
        return response.json()

    # Dropbox reports an existing folder as a 409 path/conflict/folder error.
    if response.status_code == 409:
        try:
            error_summary = response.json().get("error_summary", "")
        except ValueError:
            error_summary = ""
        if error_summary.startswith("path/conflict/folder"):
            return None

    _dropbox_error(response, f"create folder {remote_folder}")


def dropbox_move_file(access_token, from_path, to_path):
    """Move an Incoming Dropbox file to its processing destination."""
    response = requests.post(
        f"{DROPBOX_API_BASE_URL}/files/move_v2",
        headers=_dropbox_api_headers(access_token),
        json={
            "from_path": from_path,
            "to_path": to_path,
            "allow_shared_folder": False,
            "autorename": True,
            "allow_ownership_transfer": False,
        },
        timeout=DROPBOX_TIMEOUT_SECONDS,
    )
    _dropbox_error(response, f"move {from_path} to {to_path}")
    return response.json()


def _local_route_for_file(staging_root, filename):
    """Return the processing destination selected in local staging."""
    routes = [
        ("Processed", RUN_MONTH_FOLDER),
        *((folder,) for folder in DROPBOX_ROUTE_FOLDERS if folder != "Processed"),
    ]
    for route_parts in routes:
        candidate = os.path.join(staging_root, *route_parts, filename)
        if os.path.exists(candidate):
            return "/".join(route_parts)
    return None


def _print_dry_run_preview(staging_root, remote_files):
    """Print proposed import rows and routes without changing Dropbox."""
    print()
    print("=" * 65)
    print("  DRY RUN PREVIEW — DROPBOX WAS NOT CHANGED")
    print("=" * 65)

    import_rows = []
    for output_name in sorted(os.listdir(OUTPUT_FOLDER)):
        if not output_name.startswith("FP_Import_") or not output_name.endswith(".csv"):
            continue
        output_path = os.path.join(OUTPUT_FOLDER, output_name)
        with open(output_path, newline="", encoding="utf-8") as input_file:
            import_rows.extend(csv.DictReader(input_file))

    if import_rows:
        for row_number, row in enumerate(import_rows, start=1):
            print(f"\nProposed FP import row {row_number}:")
            for column, value in row.items():
                print(f"  {column}: {value}")
    else:
        print("\nNo FP import row was generated.")

    print("\nProposed Dropbox routing:")
    preview_results = []
    unused_rows = list(import_rows)
    for entry in remote_files:
        filename = entry["name"]
        route = _local_route_for_file(staging_root, filename)
        matched_row = None

        # Associate each generated row with its source invoice. Reference
        # numbers are part of the required filename, making this deterministic
        # for normal inputs. The ordered fallback covers unusual punctuation.
        for row in unused_rows:
            reference = str(row.get("Reference Number", "")).strip()
            if reference and reference.casefold() in filename.casefold():
                matched_row = row
                break
        if matched_row is None and route == "Processed" and unused_rows:
            matched_row = unused_rows[0]
        if matched_row is not None:
            unused_rows.remove(matched_row)

        print(f"  {filename} → {route or 'Remain in Incoming'}")
        preview_results.append({
            "filename": filename,
            "route": route or "Remain in Incoming",
            "row": matched_row,
        })

    print("\nNo Dropbox outputs were uploaded and no source files were moved.")
    # Machine-readable output used by the web API. Keeping it on one line
    # avoids changing the human-readable CLI preview above.
    print("ZGM_PREVIEW_JSON=" + json.dumps(preview_results, ensure_ascii=False))


def process_dropbox_receipts(selected_filenames=None, dry_run=False):
    """
    Process Dropbox Incoming files through a temporary local staging area.

    Generated outputs are uploaded before any source invoice is moved remotely.
    This keeps the Dropbox Incoming copy intact if processing or upload fails.
    """
    remote_root = os.getenv("DROPBOX_MEDIA_ROOT", "/Automation Testing")
    remote_root = remote_root.strip().strip('"')
    remote_incoming = _dropbox_remote_path(remote_root, "Incoming")

    print("=" * 65)
    print("  Dropbox Media Receipt Processing")
    print(f"  Remote root: {remote_root}")
    print("=" * 65)

    try:
        access_token = get_dropbox_access_token()
        entries = dropbox_list_folder(access_token, remote_incoming)
    except DropboxIntegrationError as exc:
        print(f"\n✗ {exc}")
        print("No Dropbox files were changed.")
        return

    remote_files = [entry for entry in entries if entry.get(".tag") == "file"]
    ignored_entries = [entry for entry in entries if entry.get(".tag") != "file"]
    for entry in ignored_entries:
        print(f"  ⚠ Ignoring non-file entry in Incoming: {entry.get('name')}")

    if selected_filenames:
        requested = {name.casefold(): name for name in selected_filenames}
        matches_by_name = {
            entry.get("name", "").casefold(): entry for entry in remote_files
        }
        missing = [
            original for normalized, original in requested.items()
            if normalized not in matches_by_name
        ]
        if missing:
            print(
                "\n✗ Dropbox Incoming does not contain the selected file(s):"
            )
            for filename in missing:
                print(f"  {filename}")
            print("No Dropbox files were changed.")
            return
        # Preserve the selection order sent by the UI.
        remote_files = [
            matches_by_name[name.casefold()] for name in selected_filenames
        ]

    if not remote_files:
        print("\nNo files found in Dropbox Incoming/.")
        return

    if dry_run:
        print("\nDRY RUN: Dropbox writes and moves are disabled.")

    with tempfile.TemporaryDirectory(prefix="zgm_media_receipts_") as staging_root:
        set_processing_root(staging_root)
        for folder in ["Incoming", "Output"] + DROPBOX_ROUTE_FOLDERS:
            os.makedirs(os.path.join(staging_root, folder), exist_ok=True)

        try:
            for entry in remote_files:
                local_path = os.path.join(INCOMING_FOLDER, entry["name"])
                print(f"  Downloading: {entry['name']}")
                dropbox_download_file(
                    access_token,
                    entry.get("path_lower") or entry["id"],
                    local_path,
                )

            # Existing extraction, Function Point lookup, routing and reports.
            process_receipts()

            if dry_run:
                _print_dry_run_preview(staging_root, remote_files)
                return

            # Upload every generated report before moving any source invoice.
            uploaded_outputs = []
            remote_output_folder = _dropbox_remote_path(
                remote_root, "Output", RUN_MONTH_FOLDER
            )
            dropbox_ensure_folder(
                access_token, _dropbox_remote_path(remote_root, "Output")
            )
            dropbox_ensure_folder(access_token, remote_output_folder)
            for output_name in sorted(os.listdir(OUTPUT_FOLDER)):
                local_output = os.path.join(OUTPUT_FOLDER, output_name)
                if not os.path.isfile(local_output):
                    continue
                remote_output = _dropbox_remote_path(
                    remote_output_folder, output_name
                )
                print(f"  Uploading output: {output_name}")
                metadata = dropbox_upload_file(
                    access_token, local_output, remote_output
                )
                uploaded_outputs.append(metadata)

            # Apply the local routing result to each original Dropbox file.
            moved_files = []
            unresolved_files = []
            for entry in remote_files:
                destination = _local_route_for_file(staging_root, entry["name"])

                if destination is None:
                    unresolved_files.append(entry["name"])
                    continue

                from_path = entry.get("path_lower") or entry["id"]
                remote_destination = _dropbox_remote_path(
                    remote_root, destination
                )
                if destination == f"Processed/{RUN_MONTH_FOLDER}":
                    dropbox_ensure_folder(
                        access_token,
                        _dropbox_remote_path(remote_root, "Processed"),
                    )
                dropbox_ensure_folder(access_token, remote_destination)
                to_path = _dropbox_remote_path(
                    remote_root, destination, entry["name"]
                )
                print(f"  Moving Dropbox file: {entry['name']} → {destination}/")
                dropbox_move_file(access_token, from_path, to_path)
                moved_files.append(entry["name"])

            print("\nDropbox synchronization complete.")
            print(f"  Outputs uploaded: {len(uploaded_outputs)}")
            print(f"  Incoming files moved: {len(moved_files)}")
            if unresolved_files:
                print("  Files left in Incoming (no routing result):")
                for name in unresolved_files:
                    print(f"    - {name}")
            print(
                "ZGM_OUTPUTS_JSON="
                + json.dumps(uploaded_outputs, ensure_ascii=False)
            )

        except (
            DropboxIntegrationError,
            FunctionPointServiceError,
            requests.RequestException,
        ) as exc:
            print(f"\n✗ Dropbox synchronization stopped: {exc}")
            print("Any source files not already reported as moved remain in Incoming/.")


# ─── Main Processing ──────────────────────────────────────────────────────────

def process_receipts():
    """Main entry point: process all files in Incoming/."""
    print("=" * 65)
    print("  ZGM Media Receipt Processing System")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 65)

    # Ensure output folders exist
    for folder in [PROCESSED_FOLDER, OUTPUT_FOLDER, ERROR_FOLDER,
                   MANUAL_REVIEW_FOLDER, MANUAL_PO_FOLDER,
                   MANUAL_MULTI_FOLDER, NAMING_ERRORS_FOLDER]:
        os.makedirs(folder, exist_ok=True)

    # Get files
    if not os.path.exists(INCOMING_FOLDER):
        print(f"\n✗ Incoming folder not found: {INCOMING_FOLDER}")
        return

    all_files = [f for f in os.listdir(INCOMING_FOLDER)
                 if not f.startswith('.') and os.path.isfile(os.path.join(INCOMING_FOLDER, f))]

    if not all_files:
        print("\nNo files found in Incoming/ folder.")
        print("Drop media invoices (PDFs) into the Incoming/ folder, then re-run.")
        return

    api_key = os.getenv("FP_API_KEY", "").strip()
    if not api_key:
        print("\n✗ FP_API_KEY is missing. Add it to the project .env file.")
        print("No invoices were processed or moved.")
        return

    fp_session = requests.Session()
    fp_session.headers.update({
        "Authorization": f"Bearer {api_key}",
        "Accept": "application/json",
    })
    fp_job_cache = {}

    print(f"\nFound {len(all_files)} file(s) to process...\n")

    # Load supplier codes
    fp_codes = load_supplier_codes()

    # Tracking
    fp_rows = []          # Rows for FP import
    multi_job_files = []  # Files with multiple job codes
    manual_review = []    # Files needing review
    naming_errors = []    # Files with bad names
    errors = []           # Processing errors
    processed = []        # Successfully processed
    summary_rows = []     # For the summary report

    for filename in all_files:
        filepath = os.path.join(INCOMING_FOLDER, filename)
        print(f"  Processing: {filename}")

        # ── Step 1: Validate filename ──────────────────────────────────────
        is_valid_name, name_issues = validate_filename(filename)
        if not is_valid_name:
            print(f"    ⚠ Naming issue: {'; '.join(name_issues)}")
            move_file(filepath, NAMING_ERRORS_FOLDER)
            naming_errors.append({"file": filename, "issues": "; ".join(name_issues)})
            summary_rows.append({
                "File": filename, "Status": "Naming Error",
                "Notes": "; ".join(name_issues),
                "Supplier": "", "Job": "", "Amount": "", "Date": ""
            })
            continue

        # ── Step 2: Extract PDF text ───────────────────────────────────────
        text = extract_pdf_text(filepath)
        if text is None:
            print(f"    ✗ Could not read PDF")
            move_file(filepath, ERROR_FOLDER)
            errors.append({"file": filename, "reason": "PDF read error"})
            summary_rows.append({
                "File": filename, "Status": "Error",
                "Notes": "Could not read PDF",
                "Supplier": "", "Job": "", "Amount": "", "Date": ""
            })
            continue

        # ── Step 3: Detect supplier ────────────────────────────────────────
        fp_code, supplier_display, tax_group, sup_confidence = detect_supplier(filename, text, fp_codes)

        # ── Step 4: Extract invoice number ────────────────────────────────
        invoice_num = extract_invoice_number(filename, text)

        # ── Step 5: Extract date ──────────────────────────────────────────
        expense_date, date_confidence = extract_date(filename, text)

        # ── Step 6: Extract job codes ─────────────────────────────────────
        job_codes, job_source = extract_job_codes(filename, text)

        # ── Step 7: Extract amounts ───────────────────────────────────────
        supplier_key = supplier_display.lower()
        subtotal, tax_amount, currency, amt_confidence = extract_amount(text, supplier_key)

        # ── Step 8: Determine confidence & routing ─────────────────────────
        confidences = [sup_confidence, date_confidence, amt_confidence]
        confidence_score = sum({'HIGH': 3, 'MEDIUM': 2, 'LOW': 1}.get(c, 1) for c in confidences)
        confidence_pct = round((confidence_score - 3) / 6 * 100)
        overall_confidence = f"{confidence_pct}%"

        flags = []
        if fp_code == "UNKNOWN":
            flags.append("UNKNOWN_SUPPLIER")
        if invoice_num == "N/A":
            flags.append("NO_INVOICE_NUMBER")
        if expense_date == "N/A":
            flags.append("NO_DATE")
        if subtotal == "N/A":
            flags.append("NO_AMOUNT")
        if not job_codes:
            flags.append("NO_JOB_CODE")
        if len(job_codes) > 1:
            flags.append("MULTI_JOB")

        # Single-job invoices require a confirmed Function Point expense match.
        fp_expense_match = None
        fp_lookup_error = ""
        if len(job_codes) == 1 and fp_code != "UNKNOWN":
            job_number = re.sub(r'\D', '', job_codes[0])
            try:
                fp_job = get_function_point_job(
                    job_number, fp_session, fp_job_cache
                )
                fp_expense_match = match_supplier_to_function_point_expense(
                    fp_job, fp_code, supplier_display
                )
            except FunctionPointLookupError as exc:
                fp_lookup_error = str(exc)
                flags.append("FP_LOOKUP_FAILED")

        flag_str = "; ".join(flags) if flags else ""

        print(f"    Supplier  : {supplier_display} [{fp_code}] ({sup_confidence})")
        print(f"    Invoice # : {invoice_num}")
        print(f"    Date      : {expense_date} ({date_confidence})")
        print(f"    Jobs      : {', '.join(job_codes) if job_codes else 'NONE'}")
        print(f"    Amount    : {currency} {subtotal} | Tax: {tax_amount} ({amt_confidence})")
        if fp_expense_match:
            print(f"    FP Expense: {fp_expense_match['expense_type']} "
                  f"[{fp_expense_match['expense_type_code']}]")
            print(f"    FP Group  : {fp_expense_match['service_group']}")
        if fp_lookup_error:
            print(f"    FP Lookup : {fp_lookup_error}")
        if flags:
            print(f"    ⚑ Flags   : {flag_str}")

        # ── Step 9: Route by job count ─────────────────────────────────────
        if len(job_codes) > 1:
            # Multi-job: FP can't auto-split — send to manual multi-job folder
            dest = move_file(filepath, MANUAL_MULTI_FOLDER)
            print(f"    → Multi-job: moved to Manual Enter - Multi-Job/")
            multi_job_files.append({
                "file": filename, "supplier": supplier_display,
                "fp_code": fp_code, "invoice": invoice_num,
                "date": expense_date, "jobs": ", ".join(job_codes),
                "subtotal": subtotal, "tax": tax_amount, "currency": currency
            })
            status = "Multi-Job (Manual)"
        elif flags and any(flag in flags for flag in (
                'UNKNOWN_SUPPLIER', 'NO_INVOICE_NUMBER', 'NO_DATE',
                'NO_AMOUNT', 'NO_JOB_CODE', 'FP_LOOKUP_FAILED')):
            # Needs manual review
            dest = move_file(filepath, MANUAL_REVIEW_FOLDER)
            print(f"    → Needs review: moved to Manual Review/")
            manual_review.append({
                "file": filename, "supplier": supplier_display,
                "fp_code": fp_code, "invoice": invoice_num,
                "date": expense_date, "jobs": ", ".join(job_codes),
                "subtotal": subtotal, "tax": tax_amount, "currency": currency,
                "service_group": (
                    fp_expense_match["service_group"] if fp_expense_match else ""
                ),
                "expense_type": (
                    fp_expense_match["expense_type"] if fp_expense_match else ""
                ),
                "flags": flag_str,
                "fp_lookup_error": fp_lookup_error,
            })
            status = "Manual Review"
        else:
            # Single job — generate FP import row
            job = re.sub(r'\D', '', job_codes[0]) if job_codes else ""
            description = build_description(supplier_display, invoice_num, job_codes, text)

            # FP row: all columns from template + Confidence + Flag (last 2 cols)
            row = {
                "Reference Number":  invoice_num,
                "*Supplier":         supplier_display,
                "Expense Date":      expense_date,
                "Payable Account":   FP_DEFAULTS["Payable_Account"],
                "Office":            FP_DEFAULTS["Office"],
                "Description":       description,
                "Terms":             FP_DEFAULTS["Terms"],
                "*Job":              job,
                "*Expense Type":     fp_expense_match["expense_type"],
                "Quantity":          FP_DEFAULTS["Quantity"],
                "Rate":              subtotal,
                "Billed":            FP_DEFAULTS["Billed"],
                "Markup%":           FP_DEFAULTS["Markup_Pct"],
                "Tax Group":         tax_group,
                "Service Group":     fp_expense_match["service_group"],
                "Tip":               FP_DEFAULTS["Tip"],
                "Discount":          FP_DEFAULTS["Discount"],
                "Override":          FP_DEFAULTS["Override"],
                "Manually Exported": FP_DEFAULTS["Manually_Exported"],
                "Start New Expense": FP_DEFAULTS["Start_New_Expense"],
                "Confidence":        overall_confidence,
                "Flag":              flag_str,
            }
            fp_rows.append(row)

            # Move to processed
            dest = move_file(filepath, PROCESSED_FOLDER)
            processed.append(filename)
            print(f"    ✓ Added to FP import")
            status = f"Processed ({overall_confidence} confidence)"

        summary_rows.append({
            "File": filename, "Status": status,
            "Notes": flag_str,
            "Supplier": supplier_display,
            "Job": ", ".join(job_codes),
            "Amount": f"{currency} {subtotal}",
            "Date": expense_date
        })
        print()

    # ── Output: FP Import CSV ─────────────────────────────────────────────────
    if fp_rows:
        csv_filename = os.path.join(OUTPUT_FOLDER, f"FP_Import_{TIMESTAMP}.csv")
        fp_columns = [
            "Reference Number", "*Supplier", "Expense Date", "Payable Account",
            "Office", "Description", "Terms", "*Job", "*Expense Type",
            "Quantity", "Rate", "Billed", "Markup%", "Tax Group",
            "Service Group", "Tip", "Discount", "Override",
            "Manually Exported", "Start New Expense", "Confidence", "Flag"
        ]
        with open(csv_filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fp_columns)
            writer.writeheader()
            writer.writerows(fp_rows)
        print(
            f"  ✓ FP Import CSV: Output/{RUN_MONTH_FOLDER}/"
            f"FP_Import_{TIMESTAMP}.csv"
        )
        print(f"    ({len(fp_rows)} line(s) ready for FunctionPointe import)")

    # ── Output: Multi-Job Summary ─────────────────────────────────────────────
    if multi_job_files:
        multi_csv = os.path.join(OUTPUT_FOLDER, f"MultiJob_Summary_{TIMESTAMP}.csv")
        with open(multi_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=multi_job_files[0].keys())
            writer.writeheader()
            writer.writerows(multi_job_files)
        print(
            f"  ✓ Multi-Job Summary: Output/{RUN_MONTH_FOLDER}/"
            f"MultiJob_Summary_{TIMESTAMP}.csv"
        )
        print(f"    ({len(multi_job_files)} invoice(s) need manual job splitting)")

    # ── Output: Manual Review ─────────────────────────────────────────────────
    if manual_review:
        review_csv = os.path.join(OUTPUT_FOLDER, f"ManualReview_{TIMESTAMP}.csv")
        with open(review_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=manual_review[0].keys())
            writer.writeheader()
            writer.writerows(manual_review)
        print(
            f"  ✓ Manual Review: Output/{RUN_MONTH_FOLDER}/"
            f"ManualReview_{TIMESTAMP}.csv"
        )

    # ── Output: Processing Summary XLSX ──────────────────────────────────────
    summary_xlsx = os.path.join(OUTPUT_FOLDER, f"Processing_Summary_{TIMESTAMP}.xlsx")
    _write_summary_xlsx(summary_xlsx, summary_rows, fp_rows, multi_job_files, manual_review, errors, naming_errors)
    print(
        f"  ✓ Summary Report: Output/{RUN_MONTH_FOLDER}/"
        f"Processing_Summary_{TIMESTAMP}.xlsx"
    )

    # ── Final Report ──────────────────────────────────────────────────────────
    print()
    print("=" * 65)
    print("  Processing Complete")
    print("=" * 65)
    print(f"  Total files:       {len(all_files)}")
    print(f"  → FP Import ready: {len(fp_rows)}")
    print(f"  → Multi-job (manual): {len(multi_job_files)}")
    print(f"  → Manual review:   {len(manual_review)}")
    print(f"  → Naming errors:   {len(naming_errors)}")
    print(f"  → Processing errors: {len(errors)}")
    print("=" * 65)

    if multi_job_files:
        print("\n  ⚑ Multi-job invoices in: Manual Enter - Multi-Job/")
        print("    Enter each job split manually in FunctionPointe.")
    if manual_review:
        print("\n  ⚑ Review items in: Manual Review/")
        print("    Check supplier codes and job assignments before importing.")
    print()


# ─── Summary XLSX Writer ──────────────────────────────────────────────────────

def _write_summary_xlsx(path, summary_rows, fp_rows, multi_rows, review_rows, errors, naming_errors):
    """Write a nicely formatted Excel summary workbook."""
    wb = Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1C1C1C")  # Dark
    ZGM_FILL    = PatternFill("solid", fgColor="E8F0FE")  # Light blue-grey
    GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
    YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
    RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
    WHITE_FONT  = Font(bold=True, color="FFFFFF")
    BOLD        = Font(bold=True)

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row_num, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = HEADER_FILL
            cell.font = WHITE_FONT
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin

    def style_row(ws, row_num, col_count, fill=None):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            if fill:
                cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # ── Sheet 1: Overview ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Overview"
    ws['A1'] = "ZGM Media Receipt Processing Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, color="666666")
    ws.merge_cells('A1:G1')
    ws.merge_cells('A2:G2')

    ws['A4'] = "Category"
    ws['B4'] = "Count"
    ws['C4'] = "Action Required"
    style_header(ws, 4, 3)

    stats = [
        ("Ready for FP Import",    len(fp_rows),       "Upload FP_Import_*.csv to FunctionPointe"),
        ("Multi-Job (Manual)",     len(multi_rows),     "Split by job in FP — see Manual Enter - Multi-Job/"),
        ("Manual Review",          len(review_rows),    "Verify supplier/job — see Manual Review/"),
        ("Naming Errors",          len(naming_errors),  "Rename files per convention — see Naming Errors/"),
        ("Processing Errors",      len(errors),         "Check PDFs — see Error/"),
    ]
    fills = [GREEN_FILL, YELLOW_FILL, YELLOW_FILL, RED_FILL, RED_FILL]
    for i, (cat, count, action) in enumerate(stats, start=5):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=count)
        ws.cell(row=i, column=3, value=action)
        style_row(ws, i, 3, fill=fills[i-5] if count > 0 else None)

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 55

    # ── Sheet 2: All Files ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Files")
    cols = ["File", "Status", "Supplier", "Job", "Amount", "Date", "Notes"]
    for c, col in enumerate(cols, 1):
        ws2.cell(row=1, column=c, value=col)
    style_header(ws2, 1, len(cols))

    status_fills = {
        "Processed": GREEN_FILL,
        "Multi-Job": YELLOW_FILL,
        "Manual":    YELLOW_FILL,
        "Error":     RED_FILL,
        "Naming":    RED_FILL,
    }
    for r, row in enumerate(summary_rows, start=2):
        vals = [row.get("File",""), row.get("Status",""), row.get("Supplier",""),
                row.get("Job",""), row.get("Amount",""), row.get("Date",""), row.get("Notes","")]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=r, column=c, value=v)
        # Pick fill based on status
        status = row.get("Status", "")
        fill = None
        for key, f in status_fills.items():
            if key.lower() in status.lower():
                fill = f
                break
        style_row(ws2, r, len(cols), fill)

    widths = [45, 30, 28, 20, 16, 14, 40]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[chr(64+i)].width = w

    # ── Sheet 3: FP Import Preview ────────────────────────────────────────────
    if fp_rows:
        ws3 = wb.create_sheet("FP Import Preview")
        fp_cols = list(fp_rows[0].keys())
        for c, col in enumerate(fp_cols, 1):
            ws3.cell(row=1, column=c, value=col)
        style_header(ws3, 1, len(fp_cols))
        for r, row in enumerate(fp_rows, start=2):
            for c, key in enumerate(fp_cols, 1):
                ws3.cell(row=r, column=c, value=row.get(key, ""))
            conf = row.get("Confidence", "")
            conf_val = int(conf.replace('%', '')) if conf and conf.endswith('%') else 0
            fill = GREEN_FILL if conf_val >= 67 else YELLOW_FILL if conf_val >= 34 else RED_FILL
            style_row(ws3, r, len(fp_cols), fill)
        for i in range(1, len(fp_cols)+1):
            ws3.column_dimensions[chr(64+i)].width = 18
        ws3.column_dimensions['B'].width = 12
        ws3.column_dimensions['F'].width = 45  # Description

    # ── Sheet 4: Multi-Job ────────────────────────────────────────────────────
    if multi_rows:
        ws4 = wb.create_sheet("Multi-Job")
        mcols = list(multi_rows[0].keys())
        for c, col in enumerate(mcols, 1):
            ws4.cell(row=1, column=c, value=col)
        style_header(ws4, 1, len(mcols))
        for r, row in enumerate(multi_rows, start=2):
            for c, key in enumerate(mcols, 1):
                ws4.cell(row=r, column=c, value=row.get(key, ""))
            style_row(ws4, r, len(mcols), YELLOW_FILL)
        for i in range(1, len(mcols)+1):
            ws4.column_dimensions[chr(64+i)].width = 20
        ws4.column_dimensions['A'].width = 45

    wb.save(path)


# ─── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Process ZGM media receipts from Dropbox or local folders."
    )
    parser.add_argument(
        "--local",
        action="store_true",
        help="Use the project-local folders instead of Dropbox.",
    )
    parser.add_argument(
        "--file",
        dest="selected_filenames",
        action="append",
        help=(
            "Process an exact filename from Dropbox Incoming. Repeat --file "
            "to process several selected invoices as one batch."
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview Dropbox processing without uploading or moving files.",
    )
    args = parser.parse_args()

    if args.local:
        if args.selected_filenames or args.dry_run:
            parser.error("--file and --dry-run are currently Dropbox-only options")
        set_processing_root(SCRIPT_DIR)
        process_receipts()
    else:
        process_dropbox_receipts(
            selected_filenames=args.selected_filenames,
            dry_run=args.dry_run,
        )
